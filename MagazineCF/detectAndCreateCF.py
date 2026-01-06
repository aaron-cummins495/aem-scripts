import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from urllib.parse import urlparse
import os
import pandas as pd
import cssutils

# Config
TEMPLATE_PATH = "/conf/au/settings/dam/cfm/models/magazine-article-model"
OUTPUT_XLSX = "output.xlsx"
BASE_CF_PATH = "/content/dam/au/cf/magazine-articles/migrated"
BASE_ASSET_PATH = "/content/dam/au/assets"
BASE_PAGE_PATH = "/content/au"
INPUT_FILE = "input.xlsx"
URLS_SHEET = "batch1"
ELEMENT_SELECTOR = "article[data-element='Magazine Article']"
URLS_HEADER = "urls"
CF_OUTPUT_FILE_NAME = "cf_out.xlsx"



def convert_url_to_path(url):
    parsed = urlparse(url)
    path = parsed.path
    dir_path = os.path.dirname(path)
    filename = os.path.basename(path)
    page_name, _ = os.path.splitext(filename)
    new_path = f"{BASE_CF_PATH}/{page_name}"
    return new_path

def clean_up_html(rawHtml):
    rawHtml = rawHtml.replace('/index.cfm', '/')
    rawHtml = rawHtml.replace('.cfm', '')
    rawHtml = rawHtml.replace('src="/', f'src="{BASE_ASSET_PATH}/')
    rawHtml = rawHtml.replace('href="/', f'href="{BASE_PAGE_PATH}/')
    return rawHtml

def get_page_name(url):
    path = urlparse(url).path
    filename = os.path.basename(path)
    page_name, _ = os.path.split(filename)
    return page_name

def get_profile_display(url):
    path = urlparse(url).path
    dir_path = os.path.dirname(path)
    _, display = os.path.split(dir_path)
    return display

def find_column(sheet, header_name):
    for col in range(1, sheet.max_column + 1):
        val = sheet.cell(row=1, column=col).value
        if val and str(val).strip() == header_name:
            return col
    raise ValueError(f"Column '{header_name}' not found in sheet '{sheet.title}'")

def expand_elements():
    wb = load_workbook(INPUT_FILE)
    urls_sheet_obj = wb[URLS_SHEET]

    # Output terminal text to log file as well
    log_file = open("detectAndCreateCF_log.txt", "w")
    failed_log_file = open("detectAndCreateCF_failed_log.txt", "w")

    # --- Find header columns ---
    header_row = 1
    ids_col = find_column(urls_sheet_obj, URLS_HEADER)
    # --- Read urls from input ---
    urlsToProcess = []
    for row in range(header_row + 1, urls_sheet_obj.max_row + 1):
        url_val = urls_sheet_obj.cell(row=row, column=ids_col).value
        if url_val:
            urlsToProcess.append(str(url_val).strip())

    headers = {"x-user-agent": "AU-AEM-Importer"}
    
    # --- Process URLs ---
    cfs = []
    for row_idx_place, url_val in enumerate(urlsToProcess):
        #stop after 10 for testing
        if row_idx_place > 10:
            break

        print(f"🔍 Processing URL {url_val}")
        log_file.write(f"? Processing URL {url_val}\n")

        try:
            response = requests.get(url_val, headers=headers, timeout=10)
            if response.status_code == 200:
                soup = BeautifulSoup(clean_up_html(response.text), "html.parser")

                # Select all sections with class matching the ELEMENT
                articleElement = soup.select(ELEMENT_SELECTOR)

                # If No elements found, print error, continue to next URL
                if len(articleElement) == 0:
                    print(f"⚠️ {url_val} → No '{ELEMENT_SELECTOR}' elements found")
                    log_file.write(f"! {url_val} -> No '{ELEMENT_SELECTOR}' elements found\n")
                    continue

                # If more than one element found, print error, continue to next URL
                if len(articleElement) != 1:
                    print(f"⚠️ {url_val} → Expected 1 '{ELEMENT_SELECTOR}' element, found {len(articleElement)}")
                    log_file.write(f"! {url_val} -> Expected 1 '{ELEMENT_SELECTOR}' element, found {len(articleElement)}\n")
                    continue

                articleElement = articleElement[0]

                # Get article header and process header content
                headerElement = articleElement.css.select_one("header.article-header")
                topicText = ''
                topicLink = ''
                titleText = ''
                publicationDate = ''
                if headerElement:
                    # Get article topic
                    topicElement = headerElement.css.select_one("span.channel")
                    if topicElement:
                        topicText = topicElement.text.strip()
                        topicLink = topicElement.css.select_one("a")['href'] if topicElement.css.select_one("a") else ''
                    
                    # Get article issue
                    issueElement = headerElement.css.select_one("time.issue")
                    publicationDate = issueElement['datetime'] if issueElement and issueElement.has_attr('datetime') else ''
                    # Convert issue date to YYYY-MM-DD format if possible
                    # (Assuming issue date is in format like "YYYY-MM-DD HH:MM:SS"")
                    try:
                        from datetime import datetime
                        date_obj = datetime.strptime(publicationDate, "%Y-%m-%d %H:%M:%S")
                        publicationDate = date_obj.strftime("%Y-%m-%d")
                    except ValueError:
                        pass

                    # Get article title
                    titleElement = headerElement.css.select_one("h1")
                    titleText = titleElement.text.strip() if titleElement else ''

                    # Get teaser blurb from header
                    teaserElement = headerElement.css.select_one("p.teaser")
                    teaserText = teaserElement.text.strip() if teaserElement else ''

                # Get teaser blurb from meta og:description
                teaserHeadElement = soup.select_one('meta[property="og:description"]')
                teaserHeadText = teaserHeadElement['content'] if teaserHeadElement else ''

                # Get article author
                authorText = ''
                authorElement = articleElement.css.select_one("p.credit.author") if headerElement else None
                authorText = authorElement.text.strip() if authorElement else ''
                authorText = authorText.replace('By ', '')

                # Get photo credit
                photoCreditText = ''
                photoCreditElement = articleElement.css.select_one("p.credit.photo") if headerElement else None
                photoCreditText = photoCreditElement.text.strip() if photoCreditElement else ''
                photoCreditText = photoCreditText.replace('Photo&shy;graphy by ', '')
                photoCreditText = photoCreditText.replace('Photography by ', '')

                # Get illustration credit
                illustrationCreditText = ''
                illustrationCreditElement = articleElement.css.select_one("p.credit.illustration") if headerElement else None
                illustrationCreditText = illustrationCreditElement.text.strip() if illustrationCreditElement else ''
                illustrationCreditText = illustrationCreditText.replace('Illustra&shy;tion by ', '')
                illustrationCreditText = illustrationCreditText.replace('Illustra­tion by ', '')

                # Get article image
                newsImagePath = ''
                altText = ''
                imageElement = articleElement.css.select_one("section.section-1 > figure > img")
                if imageElement and imageElement.has_attr('src'):
                    newsImagePath = imageElement['src']
                    altText = imageElement['alt'] if imageElement.has_attr('alt') else ''

                

                # Get article content html
                contentElement = articleElement.css.select_one("section.section-1")
                if contentElement:
                    # Remove unwanted elements from content
                    for unwanted in contentElement.select(':scope > figure'):
                        unwanted.decompose()
                contentHtml = contentElement.decode_contents() if contentElement else ''

                # Determine save path
                savePath = convert_url_to_path(url_val)

                cfs.append({
                    "path": savePath,   
                    "name": "articleCF",
                    "title": "articleCF",
                    "template": TEMPLATE_PATH,
                    "topic": topicText,
                    "topicLink": topicLink,
                    "news_title": titleText,
                    "teaser": teaserText if teaserText else teaserHeadText,
                    "showTeaser": "true" if teaserText else "false",
                    "author": authorText,
                    "illustrationBy": illustrationCreditText,
                    "photographyBy": photoCreditText,
                    "publicationDate": publicationDate,
                    "description": contentHtml,
                    "newsImage": newsImagePath,
                    "useImage": "true",
                    "altAsCaption": altText,
                })

            else:
                print(f"⚠️ {url_val} → HTTP {response.status_code}")
                log_file.write(f"! {url_val} -> HTTP {response.status_code}\n")
        except requests.exceptions.RequestException:
            print(f"❌ Failed to fetch {url_val}")
            log_file.write(f"X Failed to fetch {url_val}\n")
            print("----------------------------------")
            log_file.write("----------------------------------\n")
            failed_log_file.write(f"{url_val}\n")
            continue


        print(f"✅ Processed #{row_idx_place}/{len(urlsToProcess)}: {url_val}")
        log_file.write(f"O Processed #{row_idx_place}/{len(urlsToProcess)}: {url_val}\n")
        print("----------------------------------")
        log_file.write("----------------------------------\n")

    # --- Save CF Output ---
    cf_out_df = pd.DataFrame(cfs)
    cf_out_df.to_excel(CF_OUTPUT_FILE_NAME, index=False)
    print(f"✅ CF Output written to {CF_OUTPUT_FILE_NAME}")
    log_file.write(f"O CF Output written to {CF_OUTPUT_FILE_NAME}\n")
    log_file.close()
    failed_log_file.close()


if __name__ == "__main__":
    expand_elements()