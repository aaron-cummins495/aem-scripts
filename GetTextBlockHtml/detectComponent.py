import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from urllib.parse import urlparse
import os
import pandas as pd
import cssutils

# Config
TEMPLATE_PATH = "/conf/au/settings/dam/cfm/models/htmlEmbed"
OUTPUT_XLSX = "output.xlsx"
BASE_CF_PATH = "/content/dam/au/cf/html"
BASE_PAGE_PATH = "/content/au"


def convert_url_to_path(url):
    parsed = urlparse(url)
    path = parsed.path
    dir_path = os.path.dirname(path)
    filename = os.path.basename(path)
    page_name, _ = os.path.splitext(filename)
    new_path = f"{BASE_CF_PATH}{dir_path}/{page_name}"
    return new_path

def get_page_name(url):
    path = urlparse(url).path
    filename = os.path.basename(path)
    page_name, _ = os.path.splitext(filename)
    return page_name

def get_relevant_classes(css_parser, classList):
    relevant_classes = ""
    for rule in css_parser.cssRules:
        if rule.type == rule.STYLE_RULE:
            selector_text = rule.selectorText
            for class_name in classList:
                if f".{class_name}" in selector_text:
                    processed_selectors = '.html-embed-wrapper ' + selector_text.replace(",", ", .html-embed-wrapper")
                    if rule.parentRule is None:
                        relevant_classes += f"{processed_selectors} {{{rule.style.cssText}}}\n"
                    else:
                        relevant_classes += f"{rule.parentRule.cssText} {{ {processed_selectors} {{{rule.style.cssText}}} }}\n"
        elif rule.type == rule.MEDIA_RULE:
            media_relevant = ""
            for sub_rule in rule.cssRules:
                if sub_rule.type == sub_rule.STYLE_RULE:
                    selector_text = sub_rule.selectorText
                    for class_name in classList:
                        if f".{class_name}" in selector_text:
                            processed_selectors = '.html-embed-wrapper ' + selector_text.replace(",", ", .html-embed-wrapper")
                            media_relevant += f"{processed_selectors} {{{sub_rule.style.cssText}}}\n"
            if media_relevant:
                relevant_classes += f"@media {rule.media.mediaText} {{\n{media_relevant}}}\n"
    return relevant_classes

def invalidHtml(section):
    # return true if dl, dt, form, or table exist
    if section.css.select_one("dl, dt, form, table") is not None:
        return True
    # return true if multiple img or figure exist
    if len(section.css.select("img")) > 1 or len(section.css.select("figure")) > 1:
        return True
    # return true if multiple blockquotes exist
    if len(section.css.select("blockquote")) > 1:
        return True
    # return true if img parent is not figure
    for img in section.css.select("img"):
        if img.parent.name != "figure":
            return True
    return False

def expand_elements(
    input_file,
    url_sheet="batch1",
    url_header="URL",
    output_sheet_name="expanded"
):
    wb = load_workbook(input_file)
    url_sheet_obj = wb[url_sheet]

    # --- Find header columns ---
    header_row = 1
    def find_column(sheet, header_name):
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row=header_row, column=col).value
            if val and str(val).strip() == header_name:
                return col
        raise ValueError(f"Column '{header_name}' not found in sheet '{sheet.title}'")

    url_col = find_column(url_sheet_obj, url_header)

    # --- Read URLs ---
    urls = {}
    for row in range(header_row + 1, url_sheet_obj.max_row + 1):
        url_val = url_sheet_obj.cell(row=row, column=url_col).value
        if url_val:
            urls[str(url_val).strip()] = {"URL": str(url_val).strip()}

    # --- Create or replace output sheet ---
    if output_sheet_name in wb.sheetnames:
        del wb[output_sheet_name]
    out_sheet = wb.create_sheet(title=output_sheet_name)

    headers = {"x-user-agent": "AU-AEM-Importer"}

    # --- Process URLs ---
    row_idx = 2
    cfs = []
    for row_idx_place, url in enumerate(urls.values(), start=2):
        url_val = url["URL"]

        try:
            response = requests.get(url_val, headers=headers, timeout=10)
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, "html.parser")

                # soup find all data-element = 2016 Collapsible Content
                collapsibles = soup.find_all("section", {"data-element": "2016 Collapsible Content"})

                # soup find all data-element = 2016 Text Block
                text_blocks = soup.find_all("section", {"data-element": "2016 Text Block"})

                # iterate through all collapsibles and text blocks
                pageHasInvalidHtml = False
                for section in collapsibles + text_blocks:
                    if invalidHtml(section):
                        pageHasInvalidHtml = True
                        break
                
                if pageHasInvalidHtml:
                    out_sheet.cell(row=row_idx, column=1, value=url_val)
                    out_sheet.cell(row=row_idx, column=2, value="✅")
                    row_idx += 1


            else:
                print(f"⚠️ {url_val} → HTTP {response.status_code}")
        except requests.exceptions.RequestException:
            print(f"❌ Failed to fetch {url_val}")

        print(f"✅ Processed: {url_val}\n#{row_idx_place} of {len(urls)}")

    wb.save(input_file)
    print(f"\n✅ Results saved to '{output_sheet_name}' in {input_file}")


if __name__ == "__main__":
    expand_elements(
        "inputDetect.xlsx",
        url_sheet="batch1",
        url_header="URL",
        output_sheet_name="expanded",

    )
