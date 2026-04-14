import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from urllib.parse import urlparse
import os
import pandas as pd
import cssutils

# Config
TEMPLATE_PATH = "/conf/au/settings/dam/cfm/models/profiles"
OUTPUT_XLSX = "output.xlsx"
BASE_CF_PATH = "/content/dam/au/cf/profiles-migrated"
BASE_ASSET_PATH = "/content/dam/au/assets"
BASE_PAGE_PATH = "/content/au"
INPUT_FILE = "input.xlsx"
IDS_SHEET = "batch1"
ELEMENT_SELECTOR = "div.CS_Element_Custom > div.profile-full"
IDS_HEADER = "Eaglenet ID"
CF_OUTPUT_FILE_NAME = "cf_out.xlsx"



def convert_url_to_path(url):
    parsed = urlparse(url)
    path = parsed.path
    dir_path = os.path.dirname(path)
    filename = os.path.basename(path)
    page_name, _ = os.path.splitext(filename)
    new_path = f"{BASE_CF_PATH}{dir_path}/{page_name}"
    return new_path

def clean_up_html(rawHtml):
    rawHtml = rawHtml.replace('/index.cfm', '/')
    rawHtml = rawHtml.replace('.cfm', '')
    rawHtml = rawHtml.replace('src="/', f'src="{BASE_ASSET_PATH}/')
    rawHtml = rawHtml.replace('href="/', f'href="{BASE_PAGE_PATH}/')
    return rawHtml

def get_page_name(url):
    path = urlparse(url).path
    filename = os.path.basename(path)
    page_name, _ = os.path.split(filename)
    return page_name

def get_profile_display(url):
    path = urlparse(url).path
    dir_path = os.path.dirname(path)
    _, display = os.path.split(dir_path)
    return display

def find_column(sheet, header_name):
    for col in range(1, sheet.max_column + 1):
        val = sheet.cell(row=1, column=col).value
        if val and str(val).strip() == header_name:
            return col
    raise ValueError(f"Column '{header_name}' not found in sheet '{sheet.title}'")

def expand_elements():
    wb = load_workbook(INPUT_FILE)
    ids_sheet_obj = wb[IDS_SHEET]

    # Output terminal text to log file as well
    log_file = open("detectAndCreateCF_log.txt", "w")
    failed_log_file = open("detectAndCreateCF_failed_log.txt", "w")
    success_page_list = open("detectAndCreateCF_success_pages.txt", "w")

    # --- Find header columns ---
    header_row = 1
    ids_col = find_column(ids_sheet_obj, IDS_HEADER)

    # --- Read usernames from input ---
    idsToProcess = []
    for row in range(header_row + 1, ids_sheet_obj.max_row + 1):
        id_val = ids_sheet_obj.cell(row=row, column=ids_col).value
        if id_val:
            idsToProcess.append(str(id_val).strip())

    headers = {"x-user-agent": "AU-AEM-Importer"}

    # -- load profile report and create map of eaglenet ids ---
    reportWb = pd.read_excel("2025_profilerotreport.xlsx", sheet_name="2025_profilerotreport")
    eaglenetIdMap = {}
    for index, row in reportWb.iterrows():
        eaglenetIdMap[row['Eaglenet ID']] = row

    # -- load sadea list and create map of eaglenet ids ---
    sadeaWb = pd.read_excel("sadeaList.xlsx", sheet_name="Sheet1")
    sadeaMap = {}
    for index, row in sadeaWb.iterrows():
        sadeaMap[row['eaglenet_id']] = row

    # -- load new profile page list and create map of eaglenet ids to new profile pages --
    newProfilePagesWb = pd.read_excel("new_profile_pages.xlsx", sheet_name="Sheet1")
    newProfilePagesMap = {}
    for index, row in newProfilePagesWb.iterrows():
        newProfilePagesMap[row['Eaglenet ID']] = row
    
    # --- Process URLs ---
    cfs = []
    for row_idx_place, id in enumerate(idsToProcess):

        #stop after 10 for testing
        #if row_idx_place > 10:
        #    break

        inIdMap = id in eaglenetIdMap

        url_val = ''
        defaultProfilePage = eaglenetIdMap[id]['Default Profile Page'] if inIdMap else None
        if defaultProfilePage is None or not isinstance(defaultProfilePage, str) or defaultProfilePage.strip() == '':

            allProfilePages = eaglenetIdMap[id]['All Profile Pages'] if inIdMap else None
            if allProfilePages is not None and isinstance(allProfilePages, str) and allProfilePages.strip() != '':
                additional_urls_val = allProfilePages.strip().lower()
                additional_urls_val = additional_urls_val.replace('faculty:', '')
                additional_urls_val = additional_urls_val.replace('staff:', '')
                additional_urls_val = additional_urls_val.replace('student:', '')
                for additional_url_val in additional_urls_val.split('|'):
                    if additional_url_val != '':
                        url_val = additional_url_val.strip().lower()
                        break
        else:
            url_val = defaultProfilePage.strip().lower()
            
        if url_val == '':
            if id in sadeaMap:
                sadeaProfilePage = sadeaMap[id]['full url']
                if sadeaProfilePage is not None and isinstance(sadeaProfilePage, str) and sadeaProfilePage.strip() != '':
                    url_val = sadeaProfilePage.strip()

        if url_val == '':
            if id in newProfilePagesMap:
                newProfilePage = newProfilePagesMap[id]['URL']
                if newProfilePage is not None and isinstance(newProfilePage, str) and newProfilePage.strip() != '':
                    url_val = newProfilePage.strip()

        if url_val == '':
            print(f"❌ No URL found for Eaglenet ID {id}")
            log_file.write(f"X No URL found for Eaglenet ID {id}\n")
            failed_log_file.write(f"{id}\n")
            continue

        url_val = 'https://www.american.edu' + url_val if url_val.startswith('/') else url_val

        print(f"🔍 Processing Eaglenet ID {id} → {url_val}")
        log_file.write(f"? Processing Eaglenet ID {id} -> {url_val}\n")

        try:
            response = requests.get(url_val, headers=headers, timeout=10)
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, "html.parser")

                # Select all sections with class matching the ELEMENT
                profilesElement = soup.select(ELEMENT_SELECTOR)

                # If No elements found, print error, continue to next URL
                if len(profilesElement) == 0:
                    print(f"⚠️ {url_val} → No '{ELEMENT_SELECTOR}' elements found")
                    log_file.write(f"! {url_val} -> No '{ELEMENT_SELECTOR}' elements found\n")
                    failed_log_file.write(f"! {url_val} -> No '{ELEMENT_SELECTOR}' elements found\n")
                    continue

                # If more than one element found, print error, continue to next URL
                if len(profilesElement) != 1:
                    print(f"⚠️ {url_val} → Expected 1 '{ELEMENT_SELECTOR}' element, found {len(profilesElement)}")
                    log_file.write(f"! {url_val} -> Expected 1 '{ELEMENT_SELECTOR}' element, found {len(profilesElement)}\n")
                    failed_log_file.write(f"! {url_val} -> Expected 1 '{ELEMENT_SELECTOR}' element, found {len(profilesElement)}\n")
                    continue

                profilesElement = profilesElement[0]

                # Get profile display from url path
                profileDisplay = get_profile_display(url_val)

                # Determine force display value
                profileContentSection = profilesElement.css.select_one("section.profile-content")
                if profileContentSection is None:
                    print(f"⚠️ {url_val} → No 'section.profile-content' found")
                    log_file.write(f"! {url_val} -> No 'section.profile-content' found\n")
                    failed_log_file.write(f"! {url_val} -> No 'section.profile-content' found\n")
                    continue
                forceDisplay = eaglenetIdMap[id]['Force Profile'] if inIdMap else ''

                # Get Bio html
                bioElement = profilesElement.css.select_one("dd.bio-text")
                bioHtml = bioElement.decode_contents() if bioElement else ''

                # Get degrees and additional positions html
                superBioElement = profilesElement.css.select_one("dl.profile-info-bio")
                degreesHtml = ''
                additionalPositionsHtml = ''
                for item in superBioElement.css.select("dt, dd"):
                    if item.text.strip().lower() == 'degrees':
                        # append all neighboring dd until next dt
                        degreesHtmlParts = []
                        next_sibling = item.find_next_sibling()
                        while next_sibling and next_sibling.name == 'dd':
                            degreesHtmlParts.append(next_sibling.decode_contents())
                            next_sibling = next_sibling.find_next_sibling()
                        degreesHtml = '<br>'.join(degreesHtmlParts)
                    elif item.text.strip().lower() == 'additional positions at au':
                        # append all neighboring dd until next dt
                        additionalPositionsHtmlParts = []
                        next_sibling = item.find_next_sibling()
                        while next_sibling and next_sibling.name == 'dd':
                            additionalPositionsHtmlParts.append(next_sibling.decode_contents())
                            next_sibling = next_sibling.find_next_sibling()
                        additionalPositionsHtml = '<br>'.join(additionalPositionsHtmlParts)

                # Get Partnerships and Affiliations html
                partnershipsElement = profilesElement.css.select_one("section#profile-partnerships > div > ul")
                partnershipsHtml = partnershipsElement.decode_contents() if partnershipsElement else ''

                # Get Scholarly html
                scholarlyElement = profilesElement.css.select_one("section#profile-activities > div")
                # remove h2 from scholarlyHtml
                if scholarlyElement:
                    header = scholarlyElement.css.select_one("header")
                    if header:
                        header.decompose()
                scholarlyHtml = scholarlyElement.decode_contents() if scholarlyElement else ''

                """ # Get Scholarly html
                scholarlyElement = profilesElement.css.select_one("section#profile-activities > div")
                # select Scholarly entries which are in divs with class col-md-6
                scholarlyEntries = scholarlyElement.css.select("div.col-md-6 > div") if scholarlyElement else []
                scholarlyHtml = ''
                # loop through scholarly entries, append title (h3) and content (div) to scholarlyHtml
                # seperate entries with $#! to allow splitting in CF
                # seperate title and content with !#$ to allow splitting in CF
                for entry in scholarlyEntries:
                    titleElement = entry.css.select_one("h3")
                    titleText = titleElement.text.strip() if titleElement else ''
                    # remove title element from entry to avoid duplication in content
                    if titleElement:
                        titleElement.decompose()
                    contentHtml = entry.decode_contents()
                    scholarlyHtml += titleText + '!#$' + contentHtml + '$#!'
                # remove trailing $#! from scholarlyHtml
                if scholarlyHtml.endswith('$#!'):
                    scholarlyHtml = scholarlyHtml[:-3] """

                # Get contact info element
                contactInfoElement = profilesElement.css.select_one("dl.profile-contact-info")

                # Get profile name element
                profileNameElement = profilesElement.css.select_one('h1.profile-name')

                # Get name from contact info
                first_name = ''
                if profileNameElement:
                    nameElement = profileNameElement.css.select_one('span[itemprop=name]')
                    first_name = nameElement.text if nameElement else ''

                is_staff = True
                # Get For the Media block
                forTheMediaElements = profilesElement.css.select("div.profile-see-also dt")
                for element in forTheMediaElements:
                    if element.text.strip().lower() == 'for the media':
                        is_staff = False
                        break
                # TODO: Temporary for profiles not in Worday
                forceDisplay = 'Staff' if is_staff else 'Faculty'
                
                # Get faculty title from contact info
                faculty_title = ''
                staff_title = ''
                if profileNameElement:
                    facultyTitleElement = profileNameElement.css.select_one('small[itemprop=jobTitle]')
                    faculty_title = facultyTitleElement.text if facultyTitleElement and not is_staff else ''
                    staff_title = facultyTitleElement.text if facultyTitleElement and is_staff else ''

                # Get faculty dept name from contact info
                faculty_dept_name = ''
                staff_dept_name = ''
                if profileNameElement:
                    facultyDeptNameElement = profileNameElement.css.select_one('small[itemprop="worksFor affiliation memberOf"]')
                    faculty_dept_name = facultyDeptNameElement.text if facultyDeptNameElement and not is_staff else ''
                    staff_dept_name = facultyDeptNameElement.text if facultyDeptNameElement and is_staff else ''
                    facultyDeptNameElement = profileNameElement.css.select_one('small[itemprop="affiliation memberOf"]')
                    faculty_dept_name = facultyDeptNameElement.text if facultyDeptNameElement and not is_staff else ''
                    staff_dept_name = facultyDeptNameElement.text if facultyDeptNameElement and is_staff else ''

                # Get fso line 1
                fso1 = ''
                if contactInfoElement:
                    fsoLine1Element = contactInfoElement.css.select_one('dd.office1')
                    fso1 = fsoLine1Element.text if fsoLine1Element else ''

                # Get fso line 2
                fso2 = ''
                if contactInfoElement:
                    fso2Element = contactInfoElement.css.select_one('dd.office2')
                    fso2 = fso2Element.text if fso2Element else ''

                # Get fso line 3
                fso3 = ''
                if contactInfoElement:
                    fso3Element = contactInfoElement.css.select_one('dd.office3')
                    fso3 = fso3Element.text if fso3Element else ''

                # Get office hours html from last dd in contact info
                officeHoursHtml = ''
                if contactInfoElement:
                    dd_elements = contactInfoElement.css.select("dd")
                    if dd_elements:
                        officeHoursHtml = dd_elements[-1].decode_contents()

                # Get phone number from contact info
                altPhoneNumber = ''
                altPhoneType = ''
                if contactInfoElement:
                    altPhoneNumberElement = contactInfoElement.css.select_one("dd.profile-phone > a")
                    if (altPhoneNumberElement and altPhoneNumberElement.get('itemprop', '') == 'faxNumber'):
                        altPhoneNumber = altPhoneNumberElement.text.strip() if altPhoneNumberElement else ''
                        altPhoneNumberElement.decompose()  # remove phone number element to avoid duplication
                        altPhoneTypeElement = contactInfoElement.css.select_one("dd.profile-phone")
                        altPhoneType = altPhoneTypeElement.text.strip() if altPhoneTypeElement else ''

                # Get hide email value from contact info
                hideEmail = 'false'
                if contactInfoElement:
                    emailElement = contactInfoElement.css.select_one("dd.profile-email")
                    if emailElement and emailElement.text.strip() != '':
                        hideEmail = 'false'
                    else:
                        hideEmail = 'true'

                # Get hide phone value from contact info
                hidePhone = 'false'
                if contactInfoElement:
                    phoneElement = contactInfoElement.css.select_one("dd.profile-phone")
                    if phoneElement and phoneElement.text.strip() != '':
                        hidePhone = 'false'
                    else:
                        hidePhone = 'true'
                
                # Get See also links
                contactLinksHtml = ''
                seeAlsoLinks = profilesElement.css.select("div.profile-see-also > dl > dd")
                for link in seeAlsoLinks:
                    # if previous sibling dt text is 'See Also', add this link to contactLinksHtml
                    prev_dt = link.find_previous_sibling('dt')
                    if prev_dt and prev_dt.text.strip().lower() == 'see also':
                        contactLinksHtml += link.decode_contents() + '<br>'

                # Get areas of specialization
                areasOfSpecializationMigrated = ''
                areasOfSpecialization = profilesElement.css.select("div.profile-see-also > dl > dd")
                for entry in areasOfSpecialization:
                    # if previous sibling dt text is 'Areas of Specialization', add this entry to areasOfSpecializationHtml
                    prev_dt = entry.find_previous_sibling('dt')
                    if prev_dt and prev_dt.text.strip().lower() == 'areas of specialization':
                        areasOfSpecializationMigrated += entry.text + '|'
                # remove trailing | from areasOfSpecializationMigrated
                if areasOfSpecializationMigrated.endswith('|'):
                    areasOfSpecializationMigrated = areasOfSpecializationMigrated[:-1]

                # Add resume if present
                resume = BASE_ASSET_PATH + '/migrated-profile-resumes/' + profilesElement.css.select_one('div.profile-image-cv a')['href'].lstrip('/') if profilesElement.css.select_one('div.profile-image-cv a') else ''

                resume = eaglenetIdMap[id]['Resume'] if inIdMap else resume
                if resume and isinstance(resume, str) and resume.strip() != '':
                    resume = resume.strip()
                    resume = BASE_ASSET_PATH + '/migrated-profile-resumes/' + resume.lstrip('/')

                # Overwrite resume with CV if present
                cv = eaglenetIdMap[id]['CV'] if inIdMap else ''
                if cv and isinstance(cv, str) and cv.strip() != '':
                    cv = cv.strip()
                    cv = BASE_ASSET_PATH + '/migrated-profile-resumes/' + cv.lstrip('/')
                    resume = cv  # overwrite resume with CV

                # Add profile image if present and not default
                profileImage = BASE_ASSET_PATH + '/migrated-profile-images/' + profilesElement.css.select_one('div.profile-image-cv img')['src'].lstrip('/') if profilesElement.css.select_one('div.profile-image-cv img') else ''

                profileImage = eaglenetIdMap[id]['Profile Image'] if inIdMap else profileImage
                if profileImage and isinstance(profileImage, str) and profileImage.strip() != '':
                    profileImage = profileImage.strip()
                    if not profileImage.lower().endswith('/uploads/defaults/original/au_profile.jpg'):
                        profileImage = BASE_ASSET_PATH + '/migrated-profile-images/' + profileImage.lstrip('/')
                    else:
                        profileImage = '/content/dam/au/assets/global/images/au_profile.jpg'  # default image
                
                if profileImage.lower().endswith('/uploads/defaults/original/au_profile.jpg'):
                    profileImage = '/content/dam/au/assets/global/images/au_profile.jpg'  # default image

                # Add authorized admins if present
                authorizedAdmins = eaglenetIdMap[id]['Authorized Admins'] if inIdMap else ''
                authorizedAdminsString = ''
                if authorizedAdmins and isinstance(authorizedAdmins, str) and authorizedAdmins.strip() != '':
                    for admin in authorizedAdmins.split(','):
                        authorizedAdminsString += '{"authorizedAdminCMF":"(' + admin.strip() + '@american.edu)"}|'
                # remove trailing |, add closing bracket
                if len(authorizedAdminsString) > 1:
                    authorizedAdminsString = authorizedAdminsString[:-1]
                else:
                    authorizedAdminsString = ''

                savePath = BASE_CF_PATH + '/' + id
                if (len(id.strip()) >= 2):
                    # CFs save path is BASE_CF_PATH + first two chars of id + full id
                    savePath = BASE_CF_PATH + '/' + id[:2] + '/' + id

                cfs.append({
                    "path": savePath,   
                    "name": "profileCF",
                    "title": "profileCF",
                    "template": TEMPLATE_PATH,
                    "forceDisplay": forceDisplay,
                    "bio": bioHtml,
                    "degrees": degreesHtml,
                    "additionalPositions": additionalPositionsHtml,
                    "partnerships": partnershipsHtml,
                    "scholarly": scholarlyHtml,
                    "officeHours": officeHoursHtml,
                    "altPhone": altPhoneNumber,
                    "altPhoneType": altPhoneType,
                    "contactLinks": contactLinksHtml,
                    "resume": resume,
                    "photo": profileImage,
                    "defaultProfilePage": url_val.replace('https://www.american.edu', '').replace('.cfm', ''),
                    "authorizedAdminsMigrated": authorizedAdminsString,
                    "username": id,
                    "hideEmail": hideEmail,
                    "hidePhone": hidePhone,
                    "areasOfSpecializationMigrated": areasOfSpecializationMigrated,
                    "first_name": first_name,
                    "faculty_dept_name": faculty_dept_name,
                    "staff_dept_name": staff_dept_name,
                    "faculty_title": faculty_title,
                    "staff_title": staff_title,
                    "fso_line1": fso1,
                    "fso_line2": fso2,
                    "fso_phone": fso3,

                })

                print(f"✅ Processed #{row_idx_place}/{len(idsToProcess)}: {url_val}")
                log_file.write(f"O Processed #{row_idx_place}/{len(idsToProcess)}: {url_val}\n")
                success_page_list.write(f"{url_val}\n")
            else:
                print(f"⚠️ {url_val} → HTTP {response.status_code}")
                log_file.write(f"! {url_val} -> HTTP {response.status_code}\n")
                failed_log_file.write(f"! {url_val} -> HTTP {response.status_code}\n")
        except requests.exceptions.RequestException:
            print(f"❌ Failed to fetch {url_val}")
            log_file.write(f"X Failed to fetch {url_val}\n")
            failed_log_file.write(f"{url_val}\n")
            continue


        print("----------------------------------")
        log_file.write("----------------------------------\n")

    # --- Save CF Output ---
    cf_out_df = pd.DataFrame(cfs)
    cf_out_df.to_excel(CF_OUTPUT_FILE_NAME, index=False)
    print(f"✅ CF Output written to {CF_OUTPUT_FILE_NAME}")
    log_file.write(f"O CF Output written to {CF_OUTPUT_FILE_NAME}\n")
    log_file.close()
    failed_log_file.close()


if __name__ == "__main__":
    expand_elements()