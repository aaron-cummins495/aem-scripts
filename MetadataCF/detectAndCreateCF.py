import time
import random
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from urllib.parse import urlparse
import os
import pandas as pd

# Config
TEMPLATE_PATH = "/conf/au/settings/dam/cfm/models/seo-metadata"
BASE_CF_PATH = "/content/dam/au/cf/seo-metadata"
BASE_ASSET_PATH = "/content/dam/au/assets"
BASE_PAGE_PATH = "/content/au"
INPUT_FILE = "input.xlsx"
URLS_SHEET = "batch1"
URLS_HEADER = "urls"
CF_OUTPUT_FILE_NAME = "seo_cf_out_news_400-803.xlsx"

# Rate limiting
MIN_DELAY = 1
MAX_DELAY = 2


def create_driver():
    """Create a headless Chrome driver"""
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    
    driver = webdriver.Chrome(options=chrome_options)
    return driver


def convert_url_to_cf_path(url):
    """
    Convert URL to CF path following ACS Commons pattern
    https://www.american.edu/magazine/3-minutes-on-bees.cfm
    → /content/dam/au/cf/seo-metadata/magazine/3-minutes-on-bees/3-minutes-on-bees-seo
    """
    parsed = urlparse(url)
    path = parsed.path  # /magazine/3-minutes-on-bees.cfm
    
    # Remove file extension
    path_without_ext = os.path.splitext(path)[0]  # /magazine/3-minutes-on-bees
    
    # Split into parts
    path_parts = path_without_ext.strip('/').split('/')
    
    if len(path_parts) >= 2:
        section = path_parts[0]  # e.g., "magazine"
        page_name = path_parts[-1]  # e.g., "3-minutes-on-bees"
    elif len(path_parts) == 1:
        section = "root"
        page_name = path_parts[0]
    else:
        section = "unknown"
        page_name = "unknown"
    
    # ✅ CF Path format (matches ACS Commons bulk import)
    cf_path = f"{BASE_CF_PATH}/{section}/{page_name}/{page_name}-seo"
    
    return section, page_name, cf_path


def convert_asset_url(url):
    """Convert image/asset URL to AEM asset path"""
    if not url:
        return ""
    if url.startswith("http"):
        parsed = urlparse(url)
        path = parsed.path
        return f"{BASE_ASSET_PATH}{path}"
    elif url.startswith("/"):
        return f"{BASE_ASSET_PATH}{url}"
    else:
        return url


def sanitize_cell_value(value):
    """Remove illegal characters from cell values for Excel"""
    if not value:
        return ""
    
    if isinstance(value, str):
        # Remove control characters (except tab, newline, carriage return)
        # Allow only printable characters and common whitespace
        cleaned = ''.join(
            char for char in value 
            if ord(char) >= 32 or char in '\t\n\r'
        )
        # Remove any remaining null bytes
        cleaned = cleaned.replace('\x00', '')
        return cleaned.strip()
    
    return str(value)


def extract_meta_content(soup, property_name):
    """Extract meta tag content by property"""
    meta = soup.find("meta", property=property_name)
    if meta and meta.get("content"):
        return meta["content"]
    return ""


def extract_meta_name(soup, name):
    """Extract meta tag content by name"""
    meta = soup.find("meta", attrs={"name": name})
    if meta and meta.get("content"):
        return meta["content"]
    return ""


def is_404_page(soup):
    """
    Check if page is a 404 by looking at specific indicators
    Only check title and common 404 header text
    """
    # Check page title
    if soup.title:
        title = soup.title.string.lower() if soup.title.string else ""
        if "404" in title or "not found" in title:
            return True
    
    # Check for h1 with 404 message (but only in first 500 chars to avoid false positives)
    h1_tags = soup.find_all("h1", limit=2)
    for h1 in h1_tags:
        h1_text = h1.get_text().lower() if h1 else ""
        if "404" in h1_text or "page not found" in h1_text or "not found" in h1_text:
            return True
    
    return False


def extract_seo_metadata(url, driver):
    """Extract comprehensive SEO metadata using Selenium"""
    seo_data = {
        'title': '',
        'description': '',
        'keywords': '',
        'ogTitle': '',
        'ogDescription': '',
        'ogImage': '',
        'ogLocale': '',
        'ogType': '',
        'ogSiteName': '',
        'fbAppId': '',
        'twitterCard': '',
        'twitterTitle': '',
        'twitterDescription': '',
        'twitterSite': '',
        'twitterImage': '',
        'verifyV1': '',
        'issueDate': '',
        'issueName': '',
        'channelName': '',
        'channelLink': '',
        'listImage': '',
        'listImageAlt': '',
        'status': 'success',
        'error': ''
    }
    
    try:
        # Add random delay
        delay = random.uniform(MIN_DELAY, MAX_DELAY)
        print(f"   ⏳ Waiting {delay:.1f}s before request...")
        time.sleep(delay)
        
        print(f"   🌐 Loading page in browser...")
        driver.get(url)
        
        # Wait for page to fully load
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "head"))
        )
        
        print(f"   ✓ Page loaded, parsing...")
        
        # Get page source and parse
        soup = BeautifulSoup(driver.page_source, "html.parser")
        
        # Check if page is 404
        if is_404_page(soup):
            seo_data['status'] = 'skipped'
            seo_data['error'] = '404 - Page Not Found'
            return seo_data
        
        # Extract title - page title tag first, then og:title
        page_title = soup.title.string if soup.title else ""
        seo_data['title'] = page_title or extract_meta_content(soup, "og:title") or ""
        
        # Extract description - regular meta description first, then og:description
        seo_data['description'] = extract_meta_name(soup, "description") or extract_meta_content(soup, "og:description") or ""
        
        # Extract keywords
        seo_data['keywords'] = extract_meta_name(soup, "keywords") or ""
        
        # Extract Open Graph
        seo_data['ogTitle'] = extract_meta_content(soup, "og:title") or seo_data['title']
        seo_data['ogDescription'] = extract_meta_content(soup, "og:description") or seo_data['description']
        seo_data['ogImage'] = convert_asset_url(extract_meta_content(soup, "og:image"))
        seo_data['ogLocale'] = extract_meta_content(soup, "og:locale") or ""
        seo_data['ogType'] = extract_meta_content(soup, "og:type") or ""
        seo_data['ogSiteName'] = extract_meta_content(soup, "og:site_name") or ""
        
        # Extract Facebook - try both property and name attributes
        seo_data['fbAppId'] = (
            extract_meta_content(soup, "fb:app_id") or 
            extract_meta_name(soup, "fb:app_id") or 
            extract_meta_name(soup, "facebook:app_id") or 
            ""
        )
        
        # Extract Twitter
        seo_data['twitterCard'] = extract_meta_name(soup, "twitter:card") or ""
        seo_data['twitterTitle'] = extract_meta_name(soup, "twitter:title") or ""
        seo_data['twitterDescription'] = extract_meta_name(soup, "twitter:description") or ""
        seo_data['twitterSite'] = extract_meta_name(soup, "twitter:site") or ""
        seo_data['twitterImage'] = convert_asset_url(extract_meta_name(soup, "twitter:image"))
        
        # Extract verification
        seo_data['verifyV1'] = extract_meta_name(soup, "verify-v1") or ""
        
        # Extract Magazine/Custom
        seo_data['issueDate'] = extract_meta_name(soup, "issueDate") or ""
        seo_data['issueName'] = extract_meta_name(soup, "issueName") or ""
        seo_data['channelName'] = extract_meta_name(soup, "channelName") or ""
        seo_data['channelLink'] = extract_meta_name(soup, "channelLink") or ""
        seo_data['listImage'] = convert_asset_url(extract_meta_name(soup, "listImage")) or ""
        seo_data['listImageAlt'] = extract_meta_name(soup, "listImageAlt") or ""
        
    except Exception as e:
        seo_data['status'] = 'failed'
        seo_data['error'] = str(e)
        print(f"   ❌ Error: {str(e)[:60]}")
    
    return seo_data


def find_column(sheet, header_name):
    """Find column index by header name"""
    for col in range(1, sheet.max_column + 1):
        val = sheet.cell(row=1, column=col).value
        if val and str(val).strip() == header_name:
            return col
    raise ValueError(f"Column '{header_name}' not found in sheet '{sheet.title}'")


def scrape_seo_metadata():
    """Main function to scrape SEO metadata from URLs"""
    
    try:
        wb = load_workbook(INPUT_FILE)
        urls_sheet_obj = wb[URLS_SHEET]
    except FileNotFoundError:
        print(f"❌ Error: Input file '{INPUT_FILE}' not found")
        return
    except Exception as e:
        print(f"❌ Error loading input file: {str(e)}")
        return

    # Output logging
    log_file = open("seo_metadata_log.txt", "w")
    failed_log_file = open("seo_metadata_failed_log.txt", "w")
    skipped_log_file = open("seo_metadata_skipped_log.txt", "w")

    # --- Find header columns ---
    try:
        header_row = 1
        ids_col = find_column(urls_sheet_obj, URLS_HEADER)
    except ValueError as e:
        print(f"❌ Error: {str(e)}")
        log_file.close()
        failed_log_file.close()
        skipped_log_file.close()
        return
    
    # --- Read URLs from input ---
    urlsToProcess = []
    for row in range(header_row + 1, urls_sheet_obj.max_row + 1):
        url_val = urls_sheet_obj.cell(row=row, column=ids_col).value
        if url_val:
            url_str = str(url_val).strip()
            if not url_str.startswith(('http://', 'https://')):
                url_str = 'https://' + url_str
            urlsToProcess.append(url_str)

    print(f"\n{'='*60}")
    print(f"🚀 SEO METADATA CF GENERATOR")
    print(f"{'='*60}")
    print(f"📊 URLs to process: {len(urlsToProcess)}")
    print(f"⏱️  Delay between requests: {MIN_DELAY}-{MAX_DELAY}s")
    print(f"🌐 Using headless Chrome browser")
    print(f"📦 Model: {TEMPLATE_PATH}")
    print(f"{'='*60}\n")

    # Create Selenium driver
    print("🔧 Starting browser...")
    driver = create_driver()
    
    try:
        cfs = []
        successful_count = 0
        failed_count = 0
        skipped_count = 0
        
        for row_idx_place, url_val in enumerate(urlsToProcess):
            print(f"\n🔍 [{row_idx_place + 1}/{len(urlsToProcess)}] Processing:")
            print(f"   {url_val}")
            log_file.write(f"\n? [{row_idx_place + 1}/{len(urlsToProcess)}] Processing URL: {url_val}\n")

            section, page_name, cf_path = convert_url_to_cf_path(url_val)
            
            # Create entry with all fields pre-initialized to empty strings
            cf_entry = {
                "path": cf_path,
                "name": "pageMetadata",
                "title": "pageMetadata",
                "template": TEMPLATE_PATH,
                "description": "",
                "keywords": "",
                "ogTitle": "",
                "ogDescription": "",
                "ogImage": "",
                "ogLocale": "",
                "ogType": "",
                "ogSiteName": "",
                "fbAppId": "",
                "twitterCard": "",
                "twitterTitle": "",
                "twitterDescription": "",
                "twitterSite": "",
                "twitterImage": "",
                "verifyV1": "",
                "issueDate": "",
                "issueName": "",
                "channelName": "",
                "channelLink": "",
                "listImage": "",
                "listImageAlt": "",
            }

            # Extract SEO metadata
            seo_data = extract_seo_metadata(url_val, driver)
            
            if seo_data['status'] == 'skipped':
                # Page is 404 - skip it
                print(f"   ⏭️  SKIPPED - {seo_data['error']}")
                log_file.write(f"   - Skipped: {seo_data['error']}\n")
                skipped_log_file.write(f"{url_val} - {seo_data['error']}\n")
                skipped_count += 1
                continue  # Don't add to cfs list
                
            elif seo_data['status'] == 'success':
                # Update cf_entry with extracted data
                cf_entry["title"] = seo_data['title']
                cf_entry["description"] = seo_data['description']
                cf_entry["keywords"] = seo_data['keywords']
                cf_entry["ogTitle"] = seo_data['ogTitle']
                cf_entry["ogDescription"] = seo_data['ogDescription']
                cf_entry["ogImage"] = seo_data['ogImage']
                cf_entry["ogLocale"] = seo_data['ogLocale']
                cf_entry["ogType"] = seo_data['ogType']
                cf_entry["ogSiteName"] = seo_data['ogSiteName']
                cf_entry["fbAppId"] = seo_data['fbAppId']
                cf_entry["twitterCard"] = seo_data['twitterCard']
                cf_entry["twitterTitle"] = seo_data['twitterTitle']
                cf_entry["twitterDescription"] = seo_data['twitterDescription']
                cf_entry["twitterSite"] = seo_data['twitterSite']
                cf_entry["twitterImage"] = seo_data['twitterImage']
                cf_entry["verifyV1"] = seo_data['verifyV1']
                cf_entry["issueDate"] = seo_data['issueDate']
                cf_entry["issueName"] = seo_data['issueName']
                cf_entry["channelName"] = seo_data['channelName']
                cf_entry["channelLink"] = seo_data['channelLink']
                cf_entry["listImage"] = seo_data['listImage']
                cf_entry["listImageAlt"] = seo_data['listImageAlt']
                
                print(f"   ✅ SUCCESS - Extracted metadata")
                print(f"      📄 Section: {section}")
                print(f"      🗂️  CF Path: {cf_path}")
                log_file.write(f"   O Successfully extracted SEO metadata\n")
                log_file.write(f"      Section: {section}\n")
                log_file.write(f"      Page Name: {page_name}\n")
                log_file.write(f"      CF Path: {cf_path}\n")
                successful_count += 1
            else:
                print(f"   ❌ FAILED - {seo_data['error']}")
                log_file.write(f"   X Failed: {seo_data['error']}\n")
                failed_log_file.write(f"{url_val} - {seo_data['error']}\n")
                failed_count += 1
            
            # Always append the cf_entry (with or without metadata)
            # Sanitize all values to remove illegal Excel characters
            sanitized_entry = {key: sanitize_cell_value(val) for key, val in cf_entry.items()}
            cfs.append(sanitized_entry)

        # --- Save CF Output ---
        cf_out_df = pd.DataFrame(cfs)
        cf_out_df.to_excel(CF_OUTPUT_FILE_NAME, index=False)
        print(f"\n✅ Output written to {CF_OUTPUT_FILE_NAME}")
        print(f"   Ready for ACS Commons bulk import!")
        log_file.write(f"\nO Output written to {CF_OUTPUT_FILE_NAME}\n")
        log_file.write(f"   Ready for ACS Commons bulk import!\n")

        # --- Print Summary ---
        total_processed = len(urlsToProcess)
        success_rate = (successful_count / (total_processed - skipped_count) * 100) if (total_processed - skipped_count) > 0 else 0
        
        print(f"\n{'='*60}")
        print(f"📊 PROCESSING SUMMARY")
        print(f"{'='*60}")
        print(f"   Total processed: {total_processed}")
        print(f"   ✅ Successful: {successful_count}")
        print(f"   ❌ Failed: {failed_count}")
        print(f"   ⏭️  Skipped: {skipped_count}")
        print(f"   📈 Success Rate: {success_rate:.1f}%")
        print(f"{'='*60}\n")
        
        log_file.write(f"\n{'='*60}\n")
        log_file.write(f"SUMMARY\n")
        log_file.write(f"  Total: {total_processed}\n")
        log_file.write(f"  Successful: {successful_count}\n")
        log_file.write(f"  Failed: {failed_count}\n")
        log_file.write(f"  Skipped: {skipped_count}\n")
        log_file.write(f"  Success Rate: {success_rate:.1f}%\n")
        log_file.write(f"{'='*60}\n")
        
    finally:
        # Close browser
        print("🛑 Closing browser...")
        driver.quit()
        log_file.close()
        failed_log_file.close()
        skipped_log_file.close()


if __name__ == "__main__":
    scrape_seo_metadata()