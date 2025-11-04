import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

def expand_elements(input_file, url_sheet="batch1", element_sheet="element", url_header="URL", element_header="Component", output_sheet_name="expanded"):
    wb = load_workbook(input_file)
    url_sheet_obj = wb[url_sheet]
    element_sheet_obj = wb[element_sheet]

    # --- Find URL column
    header_row = 1
    url_col = None
    for col in range(1, url_sheet_obj.max_column + 1):
        if str(url_sheet_obj.cell(row=header_row, column=col).value).strip() == url_header:
            url_col = col
            break
    if url_col is None:
        raise ValueError(f"Column '{url_header}' not found in sheet '{url_sheet}'")

    # --- Read URLs
    urls = {}
    for row in range(header_row + 1, url_sheet_obj.max_row + 1):
        url_val = url_sheet_obj.cell(row=row, column=url_col).value
        if url_val:
            urls[str(url_val).strip()] = {"URL": str(url_val).strip()}

    # --- Find Component column
    element_col = None
    for col in range(1, element_sheet_obj.max_column + 1):
        if str(element_sheet_obj.cell(row=header_row, column=col).value).strip() == element_header:
            element_col = col
            break
    if element_col is None:
        raise ValueError(f"Column '{element_header}' not found in sheet '{element_sheet}'")

    # --- Collect components
    components = set()
    url_component_map = {}
    for row in range(header_row + 1, element_sheet_obj.max_row + 1):
        url_val = element_sheet_obj.cell(row=row, column=1).value
        comp_val = element_sheet_obj.cell(row=row, column=element_col).value
        if url_val:
            url_str = str(url_val).strip()
            url_component_map[url_str] = comp_val
            if comp_val:
                comps = [c.strip() for c in str(comp_val).split(",")]
                components.update(comps)
    components = sorted([c for c in components if c])

    # --- Create or replace output sheet
    if output_sheet_name in wb.sheetnames:
        del wb[output_sheet_name]
    out_sheet = wb.create_sheet(title=output_sheet_name)

    headers_row = ["URL", element_header] + components
    for col, header in enumerate(headers_row, 1):
        out_sheet.cell(row=1, column=col, value=header)

    # --- Custom headers for requests
    headers = {
        "x-user-agent": "AU-AEM-Importer"
    }

    # --- Process URLs
    for row_idx, url in enumerate(urls.values(), start=2):
        url_val = url["URL"]
        comp_val = url_component_map.get(url_val, "")
        out_sheet.cell(row=row_idx, column=1, value=url_val)
        out_sheet.cell(row=row_idx, column=2, value=comp_val)

        dom_matches = {}
        try:
            response = requests.get(url_val, headers=headers, timeout=10)
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, "html.parser")

                for comp in components:
                    comp_clean = comp.strip().lower()
                    found = False

                    # --- Loop through all <section> tags
                    for section in soup.find_all("section"):
                        data_element = section.get("data-element", "").strip().lower()
                        class_list = [cls.strip().lower() for cls in section.get("class", [])]

                        # --- Exact match for data-element AND hero-image-full in class
                        if data_element == comp_clean and "hero-image-full" in class_list:
                            found = True
                            break

                    dom_matches[comp] = 1 if found else 0
            else:
                for comp in components:
                    dom_matches[comp] = 0
        except requests.exceptions.RequestException:
            for comp in components:
                dom_matches[comp] = 0

        # --- Write results
        for i, comp in enumerate(components, start=3):
            out_sheet.cell(row=row_idx, column=i, value=dom_matches.get(comp, 0))

    wb.save(input_file)

if __name__ == "__main__":
    expand_elements(
        "input.xlsx",
        url_sheet="batch1",
        element_sheet="element",
        url_header="URL",
        element_header="Component",
        output_sheet_name="expanded"
    )
