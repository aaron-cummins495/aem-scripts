import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

def expand_elements(
    input_file,
    url_sheet="batch1",
    element_sheet="element",
    url_header="URL",
    element_header="Component",
    output_sheet_name="expanded"
):
    wb = load_workbook(input_file)
    url_sheet_obj = wb[url_sheet]
    element_sheet_obj = wb[element_sheet]

    # --- Find header columns ---
    header_row = 1
    def find_column(sheet, header_name):
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row=header_row, column=col).value
            if val and str(val).strip() == header_name:
                return col
        raise ValueError(f"Column '{header_name}' not found in sheet '{sheet.title}'")

    url_col = find_column(url_sheet_obj, url_header)
    element_col = find_column(element_sheet_obj, element_header)

    # --- Read URLs ---
    urls = {}
    for row in range(header_row + 1, url_sheet_obj.max_row + 1):
        url_val = url_sheet_obj.cell(row=row, column=url_col).value
        if url_val:
            urls[str(url_val).strip()] = {"URL": str(url_val).strip()}

    # --- Collect components from element sheet ---
    components = set()
    url_component_map = {}
    for row in range(header_row + 1, element_sheet_obj.max_row + 1):
        url_val = element_sheet_obj.cell(row=row, column=1).value
        comp_val = element_sheet_obj.cell(row=row, column=element_col).value
        if url_val:
            url_str = str(url_val).strip()
            comps = [c.strip() for c in str(comp_val).split(",")] if comp_val else []
            url_component_map[url_str] = comps
            components.update(comps)
    components = sorted([c for c in components if c])

    # --- Create or replace output sheet ---
    if output_sheet_name in wb.sheetnames:
        del wb[output_sheet_name]
    out_sheet = wb.create_sheet(title=output_sheet_name)

    headers_row = ["URL", element_header] + components
    for col, header in enumerate(headers_row, 1):
        out_sheet.cell(row=1, column=col, value=header)

    headers = {"x-user-agent": "AU-AEM-Importer"}

    # --- Process URLs ---
    for row_idx, url in enumerate(urls.values(), start=2):
        url_val = url["URL"]
        comps_for_url = url_component_map.get(url_val, [])
        out_sheet.cell(row=row_idx, column=1, value=url_val)
        out_sheet.cell(row=row_idx, column=2, value=", ".join(comps_for_url))

        dom_matches = {comp: 0 for comp in components}

        try:
            response = requests.get(url_val, headers=headers, timeout=10)
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, "html.parser")

                for section in soup.find_all("section"):
                    data_element = section.get("data-element", "").strip().lower()
                    class_list = [cls.strip().lower() for cls in section.get("class", [])]

                    for comp in components:
                        comp_clean = comp.lower()
                        if data_element == comp_clean or comp_clean in class_list:
                            dom_matches[comp] = 1
            else:
                print(f"⚠️ {url_val} → HTTP {response.status_code}")
        except requests.exceptions.RequestException:
            print(f"❌ Failed to fetch {url_val}")

        # --- Write results ---
        for i, comp in enumerate(components, start=3):
            out_sheet.cell(row=row_idx, column=i, value=dom_matches.get(comp, 0))

        print(f"✅ Processed: {url_val}")

    wb.save(input_file)
    print(f"\n✅ Results saved to '{output_sheet_name}' in {input_file}")


if __name__ == "__main__":
    expand_elements(
        "input.xlsx",
        url_sheet="batch1",
        element_sheet="element",
        url_header="URL",
        element_header="Component",
        output_sheet_name="expanded"
    )
