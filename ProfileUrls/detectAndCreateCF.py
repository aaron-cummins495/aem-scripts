import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from urllib.parse import urlparse
import os
import pandas as pd
import cssutils

# Config
TEMPLATE_PATH = "/conf/au/settings/dam/cfm/models/profiles"
OUTPUT_XLSX = "output.xlsx"
BASE_CF_PATH = "/content/dam/au/cf/profiles-migrated"
BASE_ASSET_PATH = "/content/dam/au/assets"
BASE_PAGE_PATH = "/content/au"
INPUT_FILE = "input.xlsx"
IDS_SHEET = "batch1"
ELEMENT_SELECTOR = "div.CS_Element_Custom > div.profile-full"
IDS_HEADER = "Eaglenet ID"
CF_OUTPUT_FILE_NAME = "cf_out.xlsx"



def convert_url_to_path(url):
    parsed = urlparse(url)
    path = parsed.path
    dir_path = os.path.dirname(path)
    filename = os.path.basename(path)
    page_name, _ = os.path.splitext(filename)
    new_path = f"{BASE_CF_PATH}{dir_path}/{page_name}"
    return new_path

def clean_up_html(rawHtml):
    rawHtml = rawHtml.replace('/index.cfm', '/')
    rawHtml = rawHtml.replace('.cfm', '')
    rawHtml = rawHtml.replace('src="/', f'src="{BASE_ASSET_PATH}/')
    rawHtml = rawHtml.replace('href="/', f'href="{BASE_PAGE_PATH}/')
    return rawHtml

def get_page_name(url):
    path = urlparse(url).path
    filename = os.path.basename(path)
    page_name, _ = os.path.split(filename)
    return page_name

def get_profile_display(url):
    path = urlparse(url).path
    dir_path = os.path.dirname(path)
    _, display = os.path.split(dir_path)
    return display

def find_column(sheet, header_name):
    for col in range(1, sheet.max_column + 1):
        val = sheet.cell(row=1, column=col).value
        if val and str(val).strip() == header_name:
            return col
    raise ValueError(f"Column '{header_name}' not found in sheet '{sheet.title}'")

def expand_elements():
    wb = load_workbook(INPUT_FILE)
    ids_sheet_obj = wb[IDS_SHEET]

    # Output terminal text to log file as well
    log_file = open("detectAndCreateCF_log.txt", "w")

    # --- Find header columns ---
    header_row = 1
    ids_col = find_column(ids_sheet_obj, IDS_HEADER)

    # --- Read usernames from input ---
    idsToProcess = []
    for row in range(header_row + 1, ids_sheet_obj.max_row + 1):
        id_val = ids_sheet_obj.cell(row=row, column=ids_col).value
        if id_val:
            idsToProcess.append(str(id_val).strip())

    headers = {"x-user-agent": "AU-AEM-Importer"}

    # -- load profile report and create map of eaglenet ids ---
    reportWb = pd.read_excel("2025_profilerotreport.xlsx", sheet_name="2025_profilerotreport")
    eaglenetIdMap = {}
    for index, row in reportWb.iterrows():
        eaglenetIdMap[row['Eaglenet ID']] = row
    
    # --- Process URLs ---
    cfs = []
    for row_idx_place, id in enumerate(idsToProcess):

        if id not in eaglenetIdMap:
            print(f"⚠️ Eaglenet ID {id} not found in report")
            log_file.write(f"! Eaglenet ID {id} not found in report\n")
            continue

        defaultProfilePage = eaglenetIdMap[id]['Default Profile Page']
        if defaultProfilePage is None or not isinstance(defaultProfilePage, str) or defaultProfilePage.strip() == '':
            print(f"! Eaglenet ID {id} has no Default Profile Page")
            log_file.write(f"! Eaglenet ID {id} has no Default Profile Page\n")
            continue

        url_val = defaultProfilePage.strip()
        url_val = 'https://www.american.edu' + url_val if url_val.startswith('/') else url_val

        print(f"🔍 Processing Eaglenet ID {id} → {url_val}")
        log_file.write(f"? Processing Eaglenet ID {id} -> {url_val}\n")
        cfs.append({
            "url": url_val,
        })

        print(f"✅ Processed #{row_idx_place}/{len(idsToProcess)}: {url_val}")
        log_file.write(f"O Processed #{row_idx_place}/{len(idsToProcess)}: {url_val}\n")
        print("----------------------------------")
        log_file.write("----------------------------------\n")

    # --- Save CF Output ---
    cf_out_df = pd.DataFrame(cfs)
    cf_out_df.to_excel(CF_OUTPUT_FILE_NAME, index=False)
    print(f"✅ CF Output written to {CF_OUTPUT_FILE_NAME}")
    log_file.write(f"O CF Output written to {CF_OUTPUT_FILE_NAME}\n")
    log_file.close()


if __name__ == "__main__":
    expand_elements()